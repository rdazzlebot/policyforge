"""Build a NIST 800-53 System Security Plan as a spreadsheet workbook.

Format: `.xlsx`. Despite the name, this is not a Microsoft-proprietary
format and needs no Excel licence — it's the open ISO/IEC 29500 (ECMA-376)
standard, written here by `openpyxl` in pure Python, and LibreOffice Calc
opens and edits it natively. It's preferred over `.ods` only because the
same file also opens unmodified in Excel and Google Sheets, and over `.csv`
because a `.csv` can't carry the dropdowns, frozen headers and per-sheet
structure that make a control table usable.

Everything here is deliberately conservative about which spreadsheet
features it uses — plain cell values, column widths, freeze panes,
autofilter, list validation bound to a named range, and simple cross-sheet
formulas. No pivot tables, rich text, or conditional formatting, all of
which are where Excel/LibreOffice fidelity actually diverges.

Structure, and where it comes from:

* **System Information** — the plan elements NIST SP 800-18 expects of an
  SSP (system identification, FIPS 199 categorization, owner, authorizing
  official, operational status, environment, interconnections).
* **Control Implementation** — one row per control: NIST's verbatim control
  text alongside the implementation status, control origination, responsible
  role and implementation narrative that the organization supplies.
* **Control Enhancements** — one row per enhancement, since enhancements are
  separately selected by baseline and separately assessed.
* **CIS Summary** — the checkbox matrix FedRAMP's SSP Appendix J "CIS
  Worksheet" uses. Its cells are formulas derived from the Control
  Implementation sheet rather than a second place to type a status, so the
  two can't drift apart.
* **Reference** — the controlled vocabularies (which also back the
  dropdowns), their definitions, and the provenance of the control data.

The status and origination vocabularies are FedRAMP's, taken from its
published SSP Appendix J CIS/CRM workbook template, because those are the
values an assessor expects to see and they're a superset of what a purely
internal SSP needs.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from policyforge.generate.policy_writer import OrgContext
from policyforge.ingest.schema import Control
from policyforge.ssp.narrative import DRAFT_PREFIX, SystemProfile

# FedRAMP SSP Appendix J, "CIS Worksheet" — Implementation Status columns.
IMPLEMENTATION_STATUS = (
    "Implemented",
    "Partially Implemented",
    "Planned",
    "Alternative Implementation",
    "N/A",
)
# FedRAMP SSP Appendix J, "CIS Worksheet" — Control Origination columns.
CONTROL_ORIGINATION = (
    "Service Provider Corporate",
    "Service Provider System Specific",
    "Service Provider Hybrid (Corporate and System Specific)",
    "Configured by Customer (Customer System Specific)",
    "Provided by Customer (Customer System Specific)",
    "Shared (Service Provider and Customer Responsibility)",
    "Inherited from Pre-Existing Authorization",
)
REVIEW_STATUS = ("Not reviewed", "Reviewed", "Needs rework")

STATUS_DEFINITIONS = {
    "Implemented": "The control is fully implemented and operating as described.",
    "Partially Implemented": "The control is partly in place; remaining work belongs in the POA&M.",
    "Planned": "The control is not yet in place; there is a documented plan and date.",
    "Alternative Implementation": (
        "A compensating control is used instead; describe it and why it is equivalent."
    ),
    "N/A": "The control does not apply to this system; justify why in the narrative.",
}

SHEET_SYSTEM = "System Information"
SHEET_CONTROLS = "Control Implementation"
SHEET_ENHANCEMENTS = "Control Enhancements"
SHEET_CIS = "CIS Summary"
SHEET_REFERENCE = "Reference"

_HEADER_FILL = "FF1F3864"
_SUBHEADER_FILL = "FFD9E2F3"
_WRAP_WIDTH = 70


def select_for_baseline(controls: list[Control], baseline: str) -> list[Control]:
    """Narrow a catalog to one 800-53 baseline (Low/Moderate/High).

    Enhancements are selected independently of their parent control, because
    the baselines select them independently: AC-2 is in Low but AC-2(1) is
    not, so a Low SSP that inherited every enhancement of every in-baseline
    control would overstate its own scope. A control is kept when the
    baseline names the control itself, and it keeps only those of its
    enhancements the baseline also names.
    """
    import dataclasses

    selected: list[Control] = []
    for control in controls:
        if baseline.lower() not in (control.baseline or "").lower():
            continue
        narrowed = dataclasses.replace(
            control,
            enhancements=[
                e for e in control.enhancements if baseline.lower() in (e.baseline or "").lower()
            ],
        )
        selected.append(narrowed)
    return selected


@dataclass
class SSPBuildResult:
    path: Path
    control_count: int
    enhancement_count: int
    narrative_count: int


def _style_header(worksheet, row: int, columns: int) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    fill = PatternFill("solid", fgColor=_HEADER_FILL)
    for column in range(1, columns + 1):
        cell = worksheet.cell(row=row, column=column)
        cell.fill = fill
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    worksheet.row_dimensions[row].height = 30


def _set_widths(worksheet, widths: dict[str, int]) -> None:
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def _wrap(worksheet, columns: list[str], first_row: int, last_row: int) -> None:
    from openpyxl.styles import Alignment

    alignment = Alignment(vertical="top", wrap_text=True)
    for column in columns:
        for row in range(first_row, last_row + 1):
            worksheet[f"{column}{row}"].alignment = alignment


def _add_list_validation(worksheet, cell_range: str, source_range: str) -> None:
    """Bind a dropdown to a range on the Reference sheet.

    A range reference is used rather than an inline comma-joined list
    because several origination values contain commas and parentheses, which
    inline list formulas quote inconsistently between Excel and LibreOffice.
    """
    from openpyxl.worksheet.datavalidation import DataValidation

    validation = DataValidation(type="list", formula1=source_range, allow_blank=True)
    validation.error = "Pick a value from the list."
    validation.errorTitle = "Invalid entry"
    worksheet.add_data_validation(validation)
    validation.add(cell_range)


def _build_reference_sheet(worksheet, catalog_version: str, generated: str) -> None:
    worksheet["A1"] = "Implementation Status"
    worksheet["B1"] = "Definition"
    worksheet["D1"] = "Control Origination"
    worksheet["F1"] = "Review Status"
    for row, status in enumerate(IMPLEMENTATION_STATUS, start=2):
        worksheet[f"A{row}"] = status
        worksheet[f"B{row}"] = STATUS_DEFINITIONS[status]
    for row, origination in enumerate(CONTROL_ORIGINATION, start=2):
        worksheet[f"D{row}"] = origination
    for row, review in enumerate(REVIEW_STATUS, start=2):
        worksheet[f"F{row}"] = review

    notes_row = max(len(IMPLEMENTATION_STATUS), len(CONTROL_ORIGINATION), len(REVIEW_STATUS)) + 4
    worksheet[f"A{notes_row}"] = "Provenance"
    worksheet[f"A{notes_row + 1}"] = "Control catalog"
    worksheet[f"B{notes_row + 1}"] = f"NIST SP 800-53 {catalog_version} (OSCAL, public domain)"
    worksheet[f"A{notes_row + 2}"] = "Catalog source"
    worksheet[f"B{notes_row + 2}"] = "https://github.com/usnistgov/oscal-content"
    worksheet[f"A{notes_row + 3}"] = "Status/origination vocabulary"
    worksheet[f"B{notes_row + 3}"] = "FedRAMP SSP Appendix J CIS/CRM Workbook"
    worksheet[f"A{notes_row + 4}"] = "Plan elements"
    worksheet[f"B{notes_row + 4}"] = "NIST SP 800-18 Rev. 1"
    worksheet[f"A{notes_row + 5}"] = "Generated"
    worksheet[f"B{notes_row + 5}"] = generated
    worksheet[f"A{notes_row + 7}"] = "Note"
    worksheet[f"B{notes_row + 7}"] = (
        "Control text is reproduced verbatim from the NIST catalog. Implementation "
        f"narratives marked '{DRAFT_PREFIX}' are machine-drafted scaffolds and must be "
        "reviewed and corrected by the system owner before this plan is relied upon."
    )
    worksheet[f"B{notes_row + 8}"] = (
        "The CIS Summary sheet is derived by formula from the Control Implementation "
        "sheet — set status and origination there, not on that sheet."
    )

    _style_header(worksheet, 1, 6)
    _set_widths(worksheet, {"A": 34, "B": 70, "C": 3, "D": 52, "E": 3, "F": 16})
    _wrap(worksheet, ["B"], 2, notes_row + 8)


def _build_system_sheet(worksheet, system: SystemProfile, org: OrgContext) -> None:
    worksheet["A1"] = "System Security Plan"
    worksheet["B1"] = system.name or "[System Name]"
    _style_header(worksheet, 1, 2)

    rows = [
        ("Organization", org.name or "[Organization Name]"),
        ("System Name", system.name),
        ("System Identifier", system.identifier),
        ("System Type", system.system_type),
        ("Operational Status", system.operational_status),
        ("", ""),
        ("FIPS 199 — Confidentiality", system.confidentiality),
        ("FIPS 199 — Integrity", system.integrity),
        ("FIPS 199 — Availability", system.availability),
        ("Overall Categorization", system.overall_categorization),
        ("", ""),
        ("System Owner", system.owner),
        ("Authorizing Official", system.authorizing_official),
        ("System Security Officer", system.security_officer),
        ("", ""),
        ("System Description / Purpose", system.description),
        ("System Environment", system.environment),
        ("System Interconnections", system.interconnections),
        (
            "Applicable Laws and Regulations",
            "; ".join(system.laws_and_regulations),
        ),
    ]

    from openpyxl.styles import Font

    for index, (label, value) in enumerate(rows, start=3):
        if not label:
            continue
        worksheet[f"A{index}"] = label
        worksheet[f"A{index}"].font = Font(bold=True)
        # An empty field becomes a visible placeholder rather than a blank
        # cell, so what still needs filling in is obvious at a glance.
        worksheet[f"B{index}"] = value or f"[{label}]"

    _set_widths(worksheet, {"A": 34, "B": 90})
    _wrap(worksheet, ["B"], 3, len(rows) + 3)


def _crosswalk_frameworks(
    controls: list[Control], crosswalk: dict[str, dict[str, list[str]]]
) -> list[str]:
    """Frameworks that actually map to at least one control in scope, so the
    workbook doesn't carry empty columns for frameworks that aren't loaded."""
    frameworks: list[str] = []
    ids = {c.control_id for c in controls}
    for control_id, mappings in crosswalk.items():
        if control_id not in ids:
            continue
        for framework in mappings:
            if framework not in frameworks:
                frameworks.append(framework)
    return sorted(frameworks)


def _build_controls_sheet(
    worksheet,
    controls: list[Control],
    narratives: dict[str, str],
    crosswalk: dict[str, dict[str, list[str]]],
    frameworks: list[str],
) -> None:
    headers = [
        "Control ID",
        "Control Name",
        "Family",
        "Baseline",
        "Control Description (NIST, verbatim)",
        "Enhancements",
        "Implementation Status",
        "Control Origination",
        "Responsible Role",
        "Implementation Description",
        "Review Status",
        "Related Controls",
    ] + [f"Maps to: {name.upper()}" for name in frameworks]

    worksheet.append(headers)
    _style_header(worksheet, 1, len(headers))

    for control in controls:
        mappings = crosswalk.get(control.control_id, {})
        row = [
            control.control_id,
            control.title,
            control.family or "",
            control.baseline or "",
            control.control_statement,
            ", ".join(e.enhancement_id for e in control.enhancements),
            "",
            "",
            "",
            narratives.get(control.control_id, ""),
            REVIEW_STATUS[0] if narratives.get(control.control_id) else "",
            ", ".join(control.related_controls),
        ] + [", ".join(mappings.get(name, [])) for name in frameworks]
        worksheet.append(row)

    last_row = worksheet.max_row
    _set_widths(
        worksheet,
        {
            "A": 12,
            "B": 34,
            "C": 22,
            "D": 20,
            "E": _WRAP_WIDTH,
            "F": 26,
            "G": 22,
            "H": 30,
            "I": 22,
            "J": _WRAP_WIDTH,
            "K": 14,
            "L": 26,
        },
    )
    _wrap(worksheet, ["E", "F", "J", "L"], 2, last_row)
    worksheet.freeze_panes = "C2"
    worksheet.auto_filter.ref = f"A1:{worksheet.cell(1, len(headers)).column_letter}{last_row}"

    if last_row > 1:
        _add_list_validation(worksheet, f"G2:G{last_row}", f"'{SHEET_REFERENCE}'!$A$2:$A$6")
        _add_list_validation(worksheet, f"H2:H{last_row}", f"'{SHEET_REFERENCE}'!$D$2:$D$8")
        _add_list_validation(worksheet, f"K2:K{last_row}", f"'{SHEET_REFERENCE}'!$F$2:$F$4")


def _build_enhancements_sheet(worksheet, controls: list[Control]) -> None:
    headers = [
        "Control ID",
        "Enhancement ID",
        "Enhancement Name",
        "Baseline",
        "Enhancement Description (NIST, verbatim)",
        "Implementation Status",
        "Control Origination",
        "Responsible Role",
        "Implementation Description",
        "Review Status",
    ]
    worksheet.append(headers)
    _style_header(worksheet, 1, len(headers))

    for control in controls:
        for enhancement in control.enhancements:
            worksheet.append(
                [
                    control.control_id,
                    enhancement.enhancement_id,
                    enhancement.title,
                    enhancement.baseline or "",
                    enhancement.description,
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
            )

    last_row = worksheet.max_row
    _set_widths(
        worksheet,
        {
            "A": 12,
            "B": 16,
            "C": 38,
            "D": 20,
            "E": _WRAP_WIDTH,
            "F": 22,
            "G": 30,
            "H": 22,
            "I": _WRAP_WIDTH,
            "J": 14,
        },
    )
    _wrap(worksheet, ["E", "I"], 2, last_row)
    worksheet.freeze_panes = "C2"
    worksheet.auto_filter.ref = f"A1:J{last_row}"

    if last_row > 1:
        _add_list_validation(worksheet, f"F2:F{last_row}", f"'{SHEET_REFERENCE}'!$A$2:$A$6")
        _add_list_validation(worksheet, f"G2:G{last_row}", f"'{SHEET_REFERENCE}'!$D$2:$D$8")
        _add_list_validation(worksheet, f"J2:J{last_row}", f"'{SHEET_REFERENCE}'!$F$2:$F$4")


def _build_cis_sheet(worksheet, controls: list[Control]) -> None:
    """FedRAMP-style CIS matrix, derived by formula from the controls sheet."""
    worksheet["A1"] = "Control Implementation Summary"
    worksheet["B1"] = (
        "Derived from the Control Implementation sheet — do not edit; "
        "change status and origination there."
    )
    worksheet.append([])

    headers = ["Control ID", *IMPLEMENTATION_STATUS, *CONTROL_ORIGINATION]
    worksheet.append(headers)
    header_row = worksheet.max_row
    _style_header(worksheet, header_row, len(headers))

    for index, control in enumerate(controls):
        source_row = index + 2  # Control Implementation data starts at row 2.
        row = [control.control_id]
        for status in IMPLEMENTATION_STATUS:
            row.append(f'=IF(\'{SHEET_CONTROLS}\'!$G{source_row}="{status}","X","")')
        for origination in CONTROL_ORIGINATION:
            row.append(f'=IF(\'{SHEET_CONTROLS}\'!$H{source_row}="{origination}","X","")')
        worksheet.append(row)

    last_row = worksheet.max_row
    _set_widths(worksheet, {"A": 12})
    for column in range(2, len(headers) + 1):
        letter = worksheet.cell(header_row, column).column_letter
        worksheet.column_dimensions[letter].width = 16
    worksheet.freeze_panes = f"B{header_row + 1}"
    if last_row > header_row:
        worksheet.auto_filter.ref = (
            f"A{header_row}:{worksheet.cell(header_row, len(headers)).column_letter}{last_row}"
        )


def build_ssp_workbook(
    controls: list[Control],
    *,
    system: SystemProfile,
    org: OrgContext,
    out_path: Path,
    crosswalk: dict[str, dict[str, list[str]]] | None = None,
    narratives: dict[str, str] | None = None,
    catalog_version: str = "Rev 5",
    generated: str = "",
) -> SSPBuildResult:
    """Write the SSP workbook and return what went into it."""
    from openpyxl import Workbook

    crosswalk = crosswalk or {}
    narratives = narratives or {}
    frameworks = _crosswalk_frameworks(controls, crosswalk)

    workbook = Workbook()
    system_sheet = workbook.active
    system_sheet.title = SHEET_SYSTEM
    _build_system_sheet(system_sheet, system, org)

    _build_controls_sheet(
        workbook.create_sheet(SHEET_CONTROLS), controls, narratives, crosswalk, frameworks
    )
    _build_enhancements_sheet(workbook.create_sheet(SHEET_ENHANCEMENTS), controls)
    _build_cis_sheet(workbook.create_sheet(SHEET_CIS), controls)
    _build_reference_sheet(workbook.create_sheet(SHEET_REFERENCE), catalog_version, generated)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out_path)

    return SSPBuildResult(
        path=out_path,
        control_count=len(controls),
        enhancement_count=sum(len(c.enhancements) for c in controls),
        narrative_count=sum(1 for v in narratives.values() if v),
    )
