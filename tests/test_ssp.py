"""ssp/ tests — baseline selection, workbook structure, and narrative drafting.

Structure assertions read the workbook back with openpyxl rather than
checking the builder's own intermediate state, so what's verified is the
file a user actually opens.

No test here calls a real LLM: `draft_implementation_narrative` is exercised
through a fake provider that records the prompt it was handed, which is what
the narrative tests are actually about — that the request is grounded in the
control and the org/system context, and that the result is safe to drop into
a spreadsheet cell.
"""

from __future__ import annotations

from pathlib import Path

import pytest

NIST_DATA = (
    Path(__file__).parent.parent / "data" / "frameworks" / "nist-800-53-r5" / "controls.json"
)
HIPAA_DATA = (
    Path(__file__).parent.parent / "data" / "frameworks" / "hipaa-security-rule" / "controls.json"
)


def _nist():
    from policyforge.ingest.schema import load_controls

    return load_controls(NIST_DATA)


def _org():
    from policyforge.generate.policy_writer import OrgContext

    return OrgContext(name="Acme Health", industry="Healthcare", vendors=["Okta", "AWS"])


def _system():
    from policyforge.ssp.narrative import SystemProfile

    return SystemProfile(
        name="Acme Health Platform",
        identifier="AHP-001",
        system_type="Major Application",
        operational_status="Operational",
        confidentiality="Moderate",
        integrity="Moderate",
        availability="Low",
        overall_categorization="Moderate",
        description="Patient scheduling and records portal.",
        laws_and_regulations=["HIPAA Security Rule", "FISMA"],
    )


class _FakeProvider:
    """Records what it was asked, returns deliberately messy multi-line text."""

    def __init__(self):
        self.calls = []

    def generate(self, *, system, prompt, max_tokens=4096, temperature=0.2):
        from policyforge.llm.base import LLMResponse

        self.calls.append({"system": system, "prompt": prompt})
        return LLMResponse(text="a. The system does\n\n   a thing.\nb. And another.", model="fake")

    def check(self):
        return True


# --------------------------------------------------------------------------
# Baseline selection
# --------------------------------------------------------------------------


@pytest.mark.parametrize("baseline,expected", [("Low", 149), ("Moderate", 287), ("High", 370)])
def test_baseline_selection_matches_nists_published_profile_totals(baseline, expected):
    """NIST's Low/Moderate/High profiles select 149/287/370 items counting
    controls and enhancements together."""
    from policyforge.ssp.workbook import select_for_baseline

    selected = select_for_baseline(_nist(), baseline)
    total = len(selected) + sum(len(c.enhancements) for c in selected)
    assert total == expected


def test_baseline_selection_filters_enhancements_independently_of_their_parent():
    """AC-2 is in Low; AC-2(1) is not. A Low plan that inherited every
    enhancement of every in-baseline control would overstate its scope."""
    from policyforge.ssp.workbook import select_for_baseline

    low = {c.control_id: c for c in select_for_baseline(_nist(), "Low")}
    moderate = {c.control_id: c for c in select_for_baseline(_nist(), "Moderate")}

    assert "AC-2" in low
    assert [e.enhancement_id for e in low["AC-2"].enhancements] == []
    assert "AC-2(1)" in {e.enhancement_id for e in moderate["AC-2"].enhancements}


def test_baseline_selection_does_not_mutate_the_source_controls():
    from policyforge.ssp.workbook import select_for_baseline

    controls = _nist()
    before = {c.control_id: len(c.enhancements) for c in controls}
    select_for_baseline(controls, "Low")
    after = {c.control_id: len(c.enhancements) for c in controls}
    assert before == after


# --------------------------------------------------------------------------
# Workbook structure
# --------------------------------------------------------------------------


def _build(tmp_path, *, narratives=None, baseline="Low"):
    from policyforge.ingest.schema import load_controls
    from policyforge.mapping.crosswalk import build_crosswalk
    from policyforge.ssp.workbook import build_ssp_workbook, select_for_baseline

    nist = _nist()
    crosswalk = build_crosswalk(nist + load_controls(HIPAA_DATA))
    scoped = select_for_baseline(nist, baseline)
    result = build_ssp_workbook(
        scoped,
        system=_system(),
        org=_org(),
        out_path=tmp_path / "ssp.xlsx",
        crosswalk=crosswalk,
        narratives=narratives or {},
        catalog_version="Rev 5 (5.2.0)",
        generated="2026-08-28",
    )
    import openpyxl

    return result, openpyxl.load_workbook(result.path), scoped


def test_workbook_has_the_expected_sheets(tmp_path):
    from policyforge.ssp import workbook as wb_mod

    _, workbook, _ = _build(tmp_path)
    assert workbook.sheetnames == [
        wb_mod.SHEET_SYSTEM,
        wb_mod.SHEET_CONTROLS,
        wb_mod.SHEET_ENHANCEMENTS,
        wb_mod.SHEET_CIS,
        wb_mod.SHEET_REFERENCE,
    ]


def test_control_sheet_carries_the_required_columns(tmp_path):
    """The columns the plan is required to contain: control, control
    description, enhancements, implementation status, implementation
    description."""
    from policyforge.ssp import workbook as wb_mod

    _, workbook, _ = _build(tmp_path)
    sheet = workbook[wb_mod.SHEET_CONTROLS]
    headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    for required in (
        "Control ID",
        "Control Description (NIST, verbatim)",
        "Enhancements",
        "Implementation Status",
        "Implementation Description",
    ):
        assert required in headers


def test_control_description_is_the_verbatim_catalog_text(tmp_path):
    """The control description must never be paraphrased or generated — it is
    NIST's authoritative wording."""
    from policyforge.ssp import workbook as wb_mod

    _, workbook, scoped = _build(tmp_path)
    sheet = workbook[wb_mod.SHEET_CONTROLS]
    by_id = {c.control_id: c for c in scoped}
    for row in range(2, sheet.max_row + 1):
        control_id = sheet.cell(row, 1).value
        assert sheet.cell(row, 5).value == (by_id[control_id].control_statement or None)


def test_status_and_origination_cells_are_dropdown_validated(tmp_path):
    """Free-typed statuses would break the CIS summary's formulas, so the
    vocabulary is enforced by list validation bound to the Reference sheet."""
    from policyforge.ssp import workbook as wb_mod

    _, workbook, _ = _build(tmp_path)
    sheet = workbook[wb_mod.SHEET_CONTROLS]
    sources = {dv.formula1 for dv in sheet.data_validations.dataValidation}
    assert f"'{wb_mod.SHEET_REFERENCE}'!$A$2:$A$6" in sources  # implementation status
    assert f"'{wb_mod.SHEET_REFERENCE}'!$D$2:$D$8" in sources  # control origination


def test_reference_sheet_lists_the_fedramp_vocabularies(tmp_path):
    from policyforge.ssp import workbook as wb_mod

    _, workbook, _ = _build(tmp_path)
    sheet = workbook[wb_mod.SHEET_REFERENCE]
    statuses = [sheet.cell(r, 1).value for r in range(2, 2 + len(wb_mod.IMPLEMENTATION_STATUS))]
    originations = [sheet.cell(r, 4).value for r in range(2, 2 + len(wb_mod.CONTROL_ORIGINATION))]
    assert statuses == list(wb_mod.IMPLEMENTATION_STATUS)
    assert originations == list(wb_mod.CONTROL_ORIGINATION)


def test_cis_summary_formulas_point_at_the_matching_control_row(tmp_path):
    """The CIS sheet is derived, not a second place to type a status. An
    off-by-one in the row reference would silently attribute one control's
    status to another."""
    from policyforge.ssp import workbook as wb_mod

    _, workbook, scoped = _build(tmp_path)
    cis = workbook[wb_mod.SHEET_CIS]
    controls_sheet = workbook[wb_mod.SHEET_CONTROLS]

    header_row = 3
    for offset in range(len(scoped)):
        cis_row = header_row + 1 + offset
        control_id = cis.cell(cis_row, 1).value
        formula = cis.cell(cis_row, 2).value
        referenced_row = int(formula.split("$G")[1].split("=")[0])
        assert controls_sheet.cell(referenced_row, 1).value == control_id


def test_crosswalk_columns_show_mapped_framework_requirements(tmp_path):
    """The point of building this inside PolicyForge: each 800-53 control
    shows the HIPAA requirements it satisfies."""
    from policyforge.ssp import workbook as wb_mod

    _, workbook, _ = _build(tmp_path)
    sheet = workbook[wb_mod.SHEET_CONTROLS]
    headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    assert "Maps to: HIPAA" in headers

    column = headers.index("Maps to: HIPAA") + 1
    rows = {
        sheet.cell(r, 1).value: sheet.cell(r, column).value for r in range(2, sheet.max_row + 1)
    }
    # AC-2 maps to the HIPAA access-management requirements via NIST's crosswalk.
    assert rows["AC-2"] and "164.308(a)(4)" in rows["AC-2"]


def test_enhancement_sheet_has_one_row_per_in_scope_enhancement(tmp_path):
    from policyforge.ssp import workbook as wb_mod

    _, workbook, scoped = _build(tmp_path)
    sheet = workbook[wb_mod.SHEET_ENHANCEMENTS]
    expected = sum(len(c.enhancements) for c in scoped)
    assert sheet.max_row - 1 == expected


def test_narratives_land_in_the_implementation_description_column(tmp_path):
    from policyforge.ssp import workbook as wb_mod
    from policyforge.ssp.narrative import DRAFT_PREFIX

    narratives = {"AC-2": f"{DRAFT_PREFIX} Accounts are managed in [Identity Provider]."}
    result, workbook, _ = _build(tmp_path, narratives=narratives)
    sheet = workbook[wb_mod.SHEET_CONTROLS]
    rows = {
        sheet.cell(r, 1).value: (sheet.cell(r, 10).value, sheet.cell(r, 11).value)
        for r in range(2, sheet.max_row + 1)
    }
    assert rows["AC-2"][0] == narratives["AC-2"]
    # A drafted row is flagged as unreviewed; a row with no narrative isn't.
    assert rows["AC-2"][1] == wb_mod.REVIEW_STATUS[0]
    assert rows["AC-1"][1] is None
    assert result.narrative_count == 1


# --------------------------------------------------------------------------
# Narrative drafting
# --------------------------------------------------------------------------


def test_narrative_prompt_is_grounded_in_the_control_and_the_org_context():
    from policyforge.ssp.narrative import draft_implementation_narrative

    control = next(c for c in _nist() if c.control_id == "AC-2")
    provider = _FakeProvider()
    draft_implementation_narrative(control, _org(), _system(), provider)

    prompt = provider.calls[0]["prompt"]
    assert "Define and document the types of accounts" in prompt  # the control's own text
    assert "Okta" in prompt  # org vendors, so it needn't invent one
    assert "Acme Health Platform" in prompt  # the system being described
    assert "Moderate" in prompt  # FIPS 199 categorization


def test_narrative_system_prompt_forbids_inventing_implementation_detail():
    """An SSP that confidently describes controls a system doesn't have is a
    false attestation, so the prompt must require placeholders over guesses."""
    from policyforge.ssp.narrative import draft_implementation_narrative

    control = next(c for c in _nist() if c.control_id == "AC-2")
    provider = _FakeProvider()
    draft_implementation_narrative(control, _org(), _system(), provider)

    system_prompt = provider.calls[0]["system"]
    assert "NEVER assert" in system_prompt
    assert "Square-Bracket Placeholder" in system_prompt


def test_narrative_is_marked_as_a_draft_and_fits_one_cell():
    from policyforge.ssp.narrative import DRAFT_PREFIX, draft_implementation_narrative

    control = next(c for c in _nist() if c.control_id == "AC-2")
    text = draft_implementation_narrative(control, _org(), _system(), _FakeProvider())

    assert text.startswith(DRAFT_PREFIX)
    # The fake returns multi-line text; newlines would break the cell layout.
    assert "\n" not in text
    assert "a. The system does a thing. b. And another." in text


def test_draft_narratives_reports_progress_per_control():
    from policyforge.ssp.narrative import draft_narratives

    controls = [c for c in _nist() if c.control_id in ("AC-1", "AC-2")]
    seen: list[str] = []
    result = draft_narratives(controls, _org(), _system(), _FakeProvider(), progress=seen.append)

    assert seen == [c.control_id for c in controls]
    assert set(result) == {"AC-1", "AC-2"}
